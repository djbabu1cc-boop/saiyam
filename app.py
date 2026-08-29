"""
Deprecito — Streamlit UI (localhost)

Upload a raw fixed-asset Excel/CSV file, optionally map columns manually, pick
reporting parameters, and download a processed Excel working with exact-day WDV
depreciation.

The UI only renders and orchestrates; ALL financial logic lives in
`depreciation_engine.py` (the `Calculate` core).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from depreciation_engine import (
    EXPECTED_COLUMNS,
    LEAP_BASIS_365,
    LEAP_BASIS_366,
    OUTPUT_COLUMNS,
    PERIOD_BASIS_FY,
    PERIOD_BASIS_CY,
    PERIOD_BASIS_QUARTER,
    PERIOD_BASIS_TILL_DATE,
    PERIOD_BASIS_OPTIONS,
    build_working,
    parse_percentage,
    working_to_dataframe,
)

st.set_page_config(
    page_title="Deprecito — Fixed Asset WDV Depreciation",
    layout="wide",
)

LEAP_OPTIONS = {
    "Leap Year (366 days)": LEAP_BASIS_366,
    "Non-Leap Year (365 days)": LEAP_BASIS_365,
}

NOT_USED = "— Not Used —"

# Fields that are truly optional (can be left unmapped)
OPTIONAL_FIELDS = {"Date of Disposal", "Sale Proceeds", "Useful Life"}


def read_uploaded_file(uploaded) -> pd.DataFrame:
    """Read any supported format (.xlsx, .xls, .xlsm, .csv, .ods).
    Always reads WITHOUT headers first. Returns DataFrame with 0-based integer
    column indices (0, 1, 2...)."""
    fname = uploaded.name.lower()
    raw_bytes = uploaded.getvalue()

    if fname.endswith(".csv"):
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(BytesIO(raw_bytes), encoding=enc, header=None)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode CSV file with any common encoding.")
    elif fname.endswith(".ods"):
        df = pd.read_excel(BytesIO(raw_bytes), engine="odf", header=None)
    else:
        df = pd.read_excel(BytesIO(raw_bytes), header=None)

    return df


def get_non_blank_columns(df: pd.DataFrame) -> list[str]:
    """Return Excel-style column letters (A, B, C...) for columns that have
    at least one non-null, non-empty value."""
    letters = []
    for idx in range(len(df.columns)):
        series = df.iloc[:, idx]
        has_data = series.notna().any() and series.dropna().astype(str).str.strip().ne("").any()
        if has_data:
            letters.append(get_column_letter(idx + 1))
    return letters


def first_row_looks_like_headers(df: pd.DataFrame) -> bool:
    """Check if row 0 of the raw data contains text that matches expected
    column headers. If so, row 0 is likely the header row."""
    if df.empty or len(df) < 1:
        return False
    first_row = df.iloc[0].astype(str).str.strip().str.lower()
    header_keywords = {
        "asset name", "gross amount", "ptu date", "date of disposal",
        "useful life", "dep rate %", "dep rate", "dep %",
        "scrap value %", "scrap %", "scrap", "sale proceeds",
    }
    matches = first_row.isin(header_keywords).sum()
    # If 3+ cells in first row match header keywords, it's likely a header row
    return matches >= 3


def apply_column_mapping(df: pd.DataFrame, mapping: dict[str, str], has_header_row: bool) -> pd.DataFrame:
    """Apply user's manual column mapping.
    mapping = {field_name: excel_letter or NOT_USED}.
    If has_header_row is True, skip row 0 (it was the header)."""
    out = pd.DataFrame()
    data_start = 1 if has_header_row else 0

    for field in EXPECTED_COLUMNS:
        letter = mapping.get(field, NOT_USED)
        if letter == NOT_USED:
            out[field] = None
        else:
            col_idx = 0
            for i, ch in enumerate(reversed(letter.upper())):
                col_idx += (ord(ch) - ord("A") + 1) * (26 ** i)
            col_idx -= 1  # Convert to 0-based
            if col_idx < len(df.columns):
                out[field] = df.iloc[data_start:, col_idx].reset_index(drop=True)
            else:
                out[field] = None

    return out


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a DataFrame with the canonical 8 columns, in spec order.
    Tries header-name matching first, then falls back to positional."""
    flat = {str(c).strip().lower(): c for c in df.columns}
    if "asset name" in flat:
        mapping = {
            "asset name": "Asset Name",
            "gross amount": "Gross Amount",
            "ptu date": "PTU Date",
            "date of disposal": "Date of Disposal",
            "useful life": "Useful Life",
            "dep rate %": "Dep Rate %",
            "scrap value %": "Scrap Value %",
            "sale proceeds": "Sale Proceeds",
        }
        alias = {
            "dep rate": "Dep Rate %",
            "dep %": "Dep Rate %",
            "scrap %": "Scrap Value %",
            "scrap": "Scrap Value %",
        }
        renamed = {}
        for key, orig in flat.items():
            canon = mapping.get(key) or alias.get(key)
            renamed[orig] = canon if canon and canon not in renamed.values() else orig
        return df.rename(columns=renamed)
    df = df.iloc[:, :8].copy()
    df.columns = EXPECTED_COLUMNS
    return df


def build_excel_bytes(out_df: pd.DataFrame, period, meta) -> bytes:
    """Render the 6-column output DataFrame as a styled .xlsx (OpenPyXL)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciation Working"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(out_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx >= 2 and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    for i, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 4)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}1"
    wb.calculation.fullCalcOnLoad = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def main() -> None:
    st.title("Deprecito")
    st.caption(
        "Fixed Asset WDV Depreciation — exact-day working file generator. "
        "Strictly offline (localhost only)."
    )

    # --- File Upload ---
    uploaded = st.file_uploader(
        "Upload the raw fixed-asset file (.xlsx, .xls, .xlsm, .csv, .ods)",
        type=["xlsx", "xls", "xlsm", "csv", "ods"],
    )

    # --- Reporting Parameters ---
    col1, col2, col3 = st.columns(3)
    with col1:
        reporting_date = st.date_input(
            "Reporting Date",
            value=date(2026, 3, 31),
        )
    with col2:
        period_basis = st.selectbox("Period Basis", PERIOD_BASIS_OPTIONS)
    with col3:
        leap_label = st.selectbox(
            "Leap Basis",
            list(LEAP_OPTIONS.keys()),
            index=1,
        )
    leap_basis = LEAP_OPTIONS[leap_label]

    till_date_basis = PERIOD_BASIS_FY
    if period_basis == PERIOD_BASIS_TILL_DATE:
        till_date_basis = st.radio(
            "Till Date Basis",
            [PERIOD_BASIS_FY, PERIOD_BASIS_CY],
            horizontal=True,
        )

    if uploaded is None:
        st.info("Upload a file to begin.")
        return

    # --- Read file (always without headers) ---
    try:
        df_raw = read_uploaded_file(uploaded)
    except Exception as exc:
        st.error(f"Could not read the file: {exc}")
        return

    if df_raw.empty:
        st.error("The uploaded file has no data.")
        return

    # --- Detect if first row is headers ---
    has_header_row = first_row_looks_like_headers(df_raw)
    data_start = 1 if has_header_row else 0

    # --- Build preview of first 3 data rows per column ---
    st.markdown("---")
    st.subheader("Column Preview")
    st.caption("First 3 data rows per column. Use this to decide your mapping.")

    preview_data = []
    for col_idx in range(len(df_raw.columns)):
        letter = get_column_letter(col_idx + 1)
        sample = df_raw.iloc[data_start:data_start + 3, col_idx].astype(str).tolist()
        sample_str = " | ".join(s if s not in ("nan", "None", "") else "(blank)" for s in sample)
        preview_data.append({"Column": letter, "Sample Data": sample_str})

    st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

    # --- Auto-detect vs Manual toggle ---
    st.markdown("---")

    # If headers detected, default to auto mode. Otherwise manual.
    use_manual_mapping = st.toggle(
        "Manual Column Mapping",
        value=not has_header_row,
        help="Turn ON to manually map columns. Turn OFF to use auto-detected headers."
    )

    mapping = {}
    if use_manual_mapping:
        st.subheader("Column Mapping")
        st.caption(
            "Map each field to the Excel column letter (A, B, C...) that contains the data. "
            "Select '— Not Used —' for optional fields that don't exist in your file."
        )

        non_blank_cols = get_non_blank_columns(df_raw)
        dropdown_options = [NOT_USED] + non_blank_cols

        # Default: positional mapping (A=Asset Name, B=Gross Amount, etc.)
        # Only for non-blank columns, up to 8 fields
        default_mapping = {}
        for i, field in enumerate(EXPECTED_COLUMNS):
            if i < len(non_blank_cols):
                default_mapping[field] = non_blank_cols[i]
            else:
                default_mapping[field] = NOT_USED

        cols_per_row = 2
        rows_needed = (len(EXPECTED_COLUMNS) + cols_per_row - 1) // cols_per_row
        for r in range(rows_needed):
            c1, c2 = st.columns(2)
            for c_idx, col_container in enumerate([c1, c2]):
                idx = r * cols_per_row + c_idx
                if idx >= len(EXPECTED_COLUMNS):
                    break
                field = EXPECTED_COLUMNS[idx]
                with col_container:
                    default_val = default_mapping.get(field, NOT_USED)
                    default_idx = dropdown_options.index(default_val) if default_val in dropdown_options else 0

                    is_optional = field in OPTIONAL_FIELDS
                    label = f"**{field}**"
                    if is_optional:
                        label += " *(optional)*"

                    selected = st.selectbox(
                        label,
                        options=dropdown_options,
                        index=default_idx,
                        key=f"map_{field}",
                    )
                    mapping[field] = selected

        # Validate mapping
        used_letters = [v for v in mapping.values() if v != NOT_USED]
        if len(set(used_letters)) != len(used_letters):
            dupes = {l for l in used_letters if used_letters.count(l) > 1}
            st.warning(
                f"⚠️ Duplicate column selection: {', '.join(sorted(dupes))}. "
                "Each field should map to a different column."
            )

        # Check required fields are mapped
        unmapped_required = [f for f in EXPECTED_COLUMNS if f not in OPTIONAL_FIELDS and mapping.get(f) == NOT_USED]
        if unmapped_required:
            st.error(
                f"❌ Required fields not mapped: {', '.join(unmapped_required)}. "
                "Please map them or the app cannot process the data."
            )

        df_work = apply_column_mapping(df_raw, mapping, has_header_row)

    else:
        # Auto mode: re-read with headers and normalize
        fname = uploaded.name.lower()
        raw_bytes = uploaded.getvalue()
        try:
            if fname.endswith(".csv"):
                for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                    try:
                        df_auto = pd.read_csv(BytesIO(raw_bytes), encoding=enc, header=0)
                        break
                    except UnicodeDecodeError:
                        continue
            elif fname.endswith(".ods"):
                df_auto = pd.read_excel(BytesIO(raw_bytes), engine="odf", header=0)
            else:
                df_auto = pd.read_excel(BytesIO(raw_bytes), header=0)
            df_work = normalize_columns(df_auto)
        except Exception as exc:
            st.error(f"Auto-detect failed: {exc}. Please toggle Manual Column Mapping.")
            return

        st.info(
            "Auto-detect mode active. First row treated as headers. "
            "Toggle 'Manual Column Mapping' above to override."
        )

    # --- Previews ---
    with st.expander("Preview mapped data", expanded=False):
        st.dataframe(df_work.head(25), width="stretch")

    # --- Generate ---
    st.markdown("---")
    if st.button("Generate Working", type="primary"):
        if df_work.empty:
            st.error("The mapped data has no rows.")
            return

        # Check required fields have data
        missing_required = [f for f in EXPECTED_COLUMNS if f not in OPTIONAL_FIELDS and df_work[f].isna().all()]
        if missing_required:
            st.error(
                f"❌ Required fields have no data: {', '.join(missing_required)}. "
                "Please check your column mapping."
            )
            return

        try:
            result = build_working(
                df_work,
                reporting_date=reporting_date,
                period_basis=period_basis,
                leap_basis=leap_basis,
                till_date_basis=till_date_basis,
            )
        except Exception as exc:
            st.error(f"Processing failed: {exc}")
            return

        out_df = working_to_dataframe(result["output"])

        st.success(
            f"Processed {len(out_df)} assets "
            f"(period {result['period'].start:%d-%m-%Y} to "
            f"{result['period'].end:%d-%m-%Y}, "
            f"{result['days_in_year']}-day basis)."
        )

        st.subheader("Output Working")
        st.dataframe(out_df, width="stretch")

        if result["excluded"]:
            with st.expander(f"Excluded assets ({len(result['excluded'])})"):
                st.dataframe(pd.DataFrame(result["excluded"]))

        if not out_df.empty:
            excel_bytes = build_excel_bytes(out_df, result["period"], result)
            st.download_button(
                label="⬇ Download Output Excel",
                data=excel_bytes,
                file_name=f"deprecito_working_{reporting_date:%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument"
                     ".spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()

st.set_page_config(
    page_title="Deprecito — Fixed Asset WDV Depreciation",
    layout="wide",
)

LEAP_OPTIONS = {
    "Leap Year (366 days)": LEAP_BASIS_366,
    "Non-Leap Year (365 days)": LEAP_BASIS_365,
}


def is_excel_letter_col(name: str) -> bool:
    """Check if a column name is an Excel letter (A, B, C... AA, AB...)."""
    if not name:
        return False
    name = str(name).strip().upper()
    if not name.isalpha():
        return False
    # Valid Excel column letters: A-Z, AA-XFD
    if len(name) == 1:
        return "A" <= name <= "Z"
    if len(name) == 2:
        return "A" <= name[0] <= "Z" and "A" <= name[1] <= "Z"
    if len(name) == 3:
        return name == "XFD" or (
            "A" <= name[0] <= "X" and "A" <= name[1] <= "Z" and "A" <= name[2] <= "Z"
        )
    return False


def all_cols_are_letters(df: pd.DataFrame) -> bool:
    """Check if ALL column names are Excel-style letters."""
    return all(is_excel_letter_col(c) for c in df.columns)


def get_non_blank_columns(df: pd.DataFrame) -> list[str]:
    """Return Excel-style column letters (A, B, C...) that have at least one
    non-null / non-empty value."""
    letters = []
    for idx, col in enumerate(df.columns):
        series = df[col]
        has_data = series.notna().any() and series.dropna().astype(str).str.strip().ne("").any()
        if has_data:
            letters.append(get_column_letter(idx + 1))
    return letters


def read_uploaded_file(uploaded) -> pd.DataFrame:
    """Read any supported format (.xlsx, .xls, .xlsm, .csv, .ods) into a
    DataFrame. Returns a DataFrame with Excel-style column letters (A, B, C...)
    as headers so mapping is position-based."""
    fname = uploaded.name.lower()
    raw_bytes = uploaded.getvalue()

    if fname.endswith(".csv"):
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(BytesIO(raw_bytes), encoding=enc, header=None)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode CSV file with any common encoding.")
    elif fname.endswith(".ods"):
        df = pd.read_excel(BytesIO(raw_bytes), engine="odf", header=None)
    else:
        df = pd.read_excel(BytesIO(raw_bytes), header=None)

    # Rename columns to A, B, C... for position-based mapping
    df.columns = [get_column_letter(i + 1) for i in range(len(df.columns))]
    return df


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a DataFrame with the canonical 8 columns, in spec order.
    Tries header-name matching first, then falls back to positional."""
    flat = {str(c).strip().lower(): c for c in df.columns}
    if "asset name" in flat:
        mapping = {
            "asset name": "Asset Name",
            "gross amount": "Gross Amount",
            "ptu date": "PTU Date",
            "date of disposal": "Date of Disposal",
            "useful life": "Useful Life",
            "dep rate %": "Dep Rate %",
            "scrap value %": "Scrap Value %",
            "sale proceeds": "Sale Proceeds",
        }
        alias = {
            "dep rate": "Dep Rate %",
            "dep %": "Dep Rate %",
            "scrap %": "Scrap Value %",
            "scrap": "Scrap Value %",
        }
        renamed = {}
        for key, orig in flat.items():
            canon = mapping.get(key) or alias.get(key)
            renamed[orig] = canon if canon and canon not in renamed.values() else orig
        return df.rename(columns=renamed)
    # No header / different headers: take the first 8 columns positionally.
    df = df.iloc[:, :8].copy()
    df.columns = EXPECTED_COLUMNS
    return df


def apply_column_mapping(df: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
    """Apply user's manual column mapping. mapping = {excel_letter: field_name}."""
    rev = {v: k for k, v in mapping.items()}
    out = pd.DataFrame()
    for field in EXPECTED_COLUMNS:
        letter = rev.get(field)
        if letter and letter in df.columns:
            out[field] = df[letter]
        else:
            out[field] = None
    return out


def build_excel_bytes(out_df: pd.DataFrame, period, meta) -> bytes:
    """Render the 6-column output DataFrame as a styled .xlsx (OpenPyXL)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciation Working"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(out_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx >= 2 and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    for i, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 4)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}1"
    wb.calculation.fullCalcOnLoad = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def main() -> None:
    st.title("Deprecito")
    st.caption(
        "Fixed Asset WDV Depreciation — exact-day working file generator. "
        "Strictly offline (localhost only)."
    )

    # --- File Upload ---
    uploaded = st.file_uploader(
        "Upload the raw fixed-asset file (.xlsx, .xls, .xlsm, .csv, .ods)",
        type=["xlsx", "xls", "xlsm", "csv", "ods"],
    )

    # --- Reporting Parameters ---
    col1, col2, col3 = st.columns(3)
    with col1:
        reporting_date = st.date_input(
            "Reporting Date",
            value=date(2026, 3, 31),
        )
    with col2:
        period_basis = st.selectbox("Period Basis", PERIOD_BASIS_OPTIONS)
    with col3:
        leap_label = st.selectbox(
            "Leap Basis",
            list(LEAP_OPTIONS.keys()),
            index=1,
        )
    leap_basis = LEAP_OPTIONS[leap_label]

    till_date_basis = PERIOD_BASIS_FY
    if period_basis == PERIOD_BASIS_TILL_DATE:
        till_date_basis = st.radio(
            "Till Date Basis",
            [PERIOD_BASIS_FY, PERIOD_BASIS_CY],
            horizontal=True,
        )

    if uploaded is None:
        st.info("Upload a file to begin.")
        return

    # --- Read file (always without headers, columns = A, B, C...) ---
    try:
        df_raw = read_uploaded_file(uploaded)
    except Exception as exc:
        st.error(f"Could not read the file: {exc}")
        return

    non_blank_cols = get_non_blank_columns(df_raw)
    if not non_blank_cols:
        st.error("No columns with data found in the uploaded file.")
        return

    # --- Smart Auto-Detect: Check if original file had real headers ---
    # Re-read with header=0 to check if first row contains header names
    fname = uploaded.name.lower()
    raw_bytes = uploaded.getvalue()
    has_real_headers = False
    try:
        if fname.endswith(".csv"):
            for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                try:
                    df_check = pd.read_csv(BytesIO(raw_bytes), encoding=enc, header=0, nrows=1)
                    break
                except UnicodeDecodeError:
                    continue
        elif fname.endswith(".ods"):
            df_check = pd.read_excel(BytesIO(raw_bytes), engine="odf", header=0, nrows=1)
        else:
            df_check = pd.read_excel(BytesIO(raw_bytes), header=0, nrows=1)

        # Check if the detected headers match our expected columns
        flat = {str(c).strip().lower(): c for c in df_check.columns}
        has_real_headers = any(
            h in flat for h in [
                "asset name", "gross amount", "ptu date", "date of disposal",
                "useful life", "dep rate %", "scrap value %", "sale proceeds",
                "dep rate", "dep %", "scrap %", "scrap",
            ]
        )
    except Exception:
        has_real_headers = False

    # --- Optional Manual Mapping Section ---
    st.markdown("---")

    use_manual_mapping = st.toggle(
        "Manual Column Mapping",
        value=not has_real_headers,
        help="Turn ON to manually map columns. Turn OFF to use auto-detected headers."
    )

    mapping = {}
    if use_manual_mapping:
        st.subheader("Column Mapping")
        st.caption(
            "Map each required field to the Excel column letter (A, B, C...) "
            "that contains the data. Only columns with actual data are shown."
        )

        # Auto-suggest mapping by position if exactly 8 non-blank cols
        auto_mapped = False
        if len(non_blank_cols) == len(EXPECTED_COLUMNS):
            for i, field in enumerate(EXPECTED_COLUMNS):
                mapping[field] = non_blank_cols[i]
            auto_mapped = True
            st.info(
                f"Auto-mapped {len(EXPECTED_COLUMNS)} fields to columns "
                f"{non_blank_cols[0]}–{non_blank_cols[-1]}. Adjust if needed."
            )

        cols_per_row = 2
        rows_needed = (len(EXPECTED_COLUMNS) + cols_per_row - 1) // cols_per_row
        for r in range(rows_needed):
            c1, c2 = st.columns(2)
            for c_idx, col_container in enumerate([c1, c2]):
                idx = r * cols_per_row + c_idx
                if idx >= len(EXPECTED_COLUMNS):
                    break
                field = EXPECTED_COLUMNS[idx]
                with col_container:
                    default_idx = 0
                    if auto_mapped and mapping.get(field) in non_blank_cols:
                        default_idx = non_blank_cols.index(mapping[field])
                    selected = st.selectbox(
                        f"**{field}**",
                        options=non_blank_cols,
                        index=default_idx,
                        key=f"map_{field}",
                    )
                    mapping[field] = selected

        # Validate mapping
        used_letters = list(mapping.values())
        if len(set(used_letters)) != len(used_letters):
            dupes = {l for l in used_letters if used_letters.count(l) > 1}
            st.warning(
                f"⚠️ Duplicate column selection: {', '.join(sorted(dupes))}. "
                "Each field should map to a different column."
            )

        df_work = apply_column_mapping(df_raw, mapping)
    else:
        # Auto mode: re-read with headers and normalize
        try:
            if fname.endswith(".csv"):
                for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                    try:
                        df_auto = pd.read_csv(BytesIO(raw_bytes), encoding=enc, header=0)
                        break
                    except UnicodeDecodeError:
                        continue
            elif fname.endswith(".ods"):
                df_auto = pd.read_excel(BytesIO(raw_bytes), engine="odf", header=0)
            else:
                df_auto = pd.read_excel(BytesIO(raw_bytes), header=0)
            df_work = normalize_columns(df_auto)
        except Exception:
            # Fallback to positional
            df_work = df_raw.iloc[:, :8].copy()
            df_work.columns = EXPECTED_COLUMNS

        st.info(
            "Auto-detect mode active. Columns mapped by header name. "
            "Toggle 'Manual Column Mapping' above to override."
        )

    # --- Previews ---
    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df_raw.head(25), width="stretch")

    with st.expander("Preview mapped data", expanded=False):
        st.dataframe(df_work.head(25), width="stretch")

    # --- Generate ---
    st.markdown("---")
    if st.button("Generate Working", type="primary"):
        if df_work.empty:
            st.error("The mapped data has no rows.")
            return

        try:
            result = build_working(
                df_work,
                reporting_date=reporting_date,
                period_basis=period_basis,
                leap_basis=leap_basis,
                till_date_basis=till_date_basis,
            )
        except Exception as exc:
            st.error(f"Processing failed: {exc}")
            return

        out_df = working_to_dataframe(result["output"])

        st.success(
            f"Processed {len(out_df)} assets "
            f"(period {result['period'].start:%d-%m-%Y} to "
            f"{result['period'].end:%d-%m-%Y}, "
            f"{result['days_in_year']}-day basis)."
        )

        st.subheader("Output Working")
        st.dataframe(out_df, width="stretch")

        if result["excluded"]:
            with st.expander(f"Excluded assets ({len(result['excluded'])})"):
                st.dataframe(pd.DataFrame(result["excluded"]))

        if not out_df.empty:
            excel_bytes = build_excel_bytes(out_df, result["period"], result)
            st.download_button(
                label="⬇ Download Output Excel",
                data=excel_bytes,
                file_name=f"deprecito_working_{reporting_date:%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument"
                     ".spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
st.set_page_config(
    page_title="Deprecito — Fixed Asset WDV Depreciation",
    layout="wide",
)

LEAP_OPTIONS = {
    "Leap Year (366 days)": LEAP_BASIS_366,
    "Non-Leap Year (365 days)": LEAP_BASIS_365,
}


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a DataFrame with the canonical 8 columns, in spec order."""
    # If an expected header is present, align by name (smart about whitespace/case).
    flat = {str(c).strip().lower(): c for c in df.columns}
    if "asset name" in flat:
        mapping = {
            "asset name": "Asset Name",
            "gross amount": "Gross Amount",
            "ptu date": "PTU Date",
            "date of disposal": "Date of Disposal",
            "useful life": "Useful Life",
            "dep rate %": "Dep Rate %",
            "scrap value %": "Scrap Value %",
            "sale proceeds": "Sale Proceeds",
        }
        # Also accept "dep rate" / "scrap %" aliases.
        alias = {
            "dep rate": "Dep Rate %",
            "dep %": "Dep Rate %",
            "scrap %": "Scrap Value %",
            "scrap": "Scrap Value %",
        }
        renamed = {}
        for key, orig in flat.items():
            canon = mapping.get(key) or alias.get(key)
            renamed[orig] = canon if canon and canon not in renamed.values() else orig
        return df.rename(columns=renamed)
    # No header / different headers: take the first 8 columns positionally.
    df = df.iloc[:, :8].copy()
    df.columns = EXPECTED_COLUMNS
    return df


def build_excel_bytes(out_df: pd.DataFrame, period, meta) -> bytes:
    """Render the 6-column output DataFrame as a styled .xlsx (OpenPyXL)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciation Working"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    money_font = Font()

    for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(out_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx >= 2 and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    # Column widths
    for i, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 4)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}1"

    # Trivial rounding guard so P/L shows 0.00 rather than tiny floats.
    wb.calculation.fullCalcOnLoad = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def main() -> None:
    st.title("Deprecito")
    st.caption(
        "Fixed Asset WDV Depreciation — exact-day working file generator. "
        "Strictly offline (localhost only)."
    )

    st.markdown(
        "**Input columns (in order):** Asset Name | Gross Amount | PTU Date | "
        "Date of Disposal | Useful Life | Dep Rate % | Scrap Value % | Sale Proceeds"
    )

    uploaded = st.file_uploader(
        "Upload the raw fixed-asset Excel file (.xlsx)",
        type=["xlsx", "xls"],
    )

    col1, col2, col3 = st.columns(3)
    with col1:
        reporting_date = st.date_input(
            "Reporting Date",
            value=date(2026, 3, 31),
        )
    with col2:
        period_basis = st.selectbox("Period Basis", PERIOD_BASIS_OPTIONS)
    with col3:
        leap_label = st.selectbox(
            "Leap Basis",
            list(LEAP_OPTIONS.keys()),
            index=1,  # default Non-Leap (365)
        )
    leap_basis = LEAP_OPTIONS[leap_label]

    till_date_basis = PERIOD_BASIS_FY
    if period_basis == PERIOD_BASIS_TILL_DATE:
        till_date_basis = st.radio(
            "Till Date Basis",
            [PERIOD_BASIS_FY, PERIOD_BASIS_CY],
            horizontal=True,
        )

    if uploaded is None:
        st.info("Upload a file to begin.")
        return

    try:
        df = pd.read_excel(uploaded)
        df = normalize_columns(df)
    except Exception as exc:
        st.error(f"Could not read the Excel file: {exc}")
        return

    # Quick preview
    with st.expander("Preview uploaded data", expanded=False):
        st.dataframe(df.head(25), width="stretch")

    st.markdown("---")
    if st.button("Generate Working", type="primary"):
        if df.empty:
            st.error("The uploaded file has no data rows.")
            return

        try:
            result = build_working(
                df,
                reporting_date=reporting_date,
                period_basis=period_basis,
                leap_basis=leap_basis,
                till_date_basis=till_date_basis,
            )
        except Exception as exc:
            st.error(f"Processing failed: {exc}")
            return

        out_df = working_to_dataframe(result["output"])

        st.success(
            f"Processed {len(out_df)} assets "
            f"(period {result['period'].start:%d-%m-%Y} to "
            f"{result['period'].end:%d-%m-%Y}, "
            f"{result['days_in_year']}-day basis)."
        )

        st.subheader("Output Working")
        st.dataframe(out_df, width="stretch")

        if result["excluded"]:
            with st.expander(
                f"Excluded assets ({len(result['excluded'])})"
            ):
                st.dataframe(pd.DataFrame(result["excluded"]))

        if not out_df.empty:
            excel_bytes = build_excel_bytes(out_df, result["period"], result)

            st.download_button(
                label="⬇ Download Output Excel",
                data=excel_bytes,
                file_name=(
                    f"deprecito_working_"
                    f"{reporting_date:%Y%m%d}.xlsx"
                ),
                mime="application/vnd.openxmlformats-officedocument"
                     ".spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
